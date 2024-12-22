import openpyxl

def merge_sheets(sheet_a_path, sheet_b_path, output_path):
    # Load Sheet A
    wb_a = openpyxl.load_workbook(sheet_a_path)
    sheet_a = wb_a.active

    # Debug: Print headers
    headers_a = [cell.value for cell in sheet_a[1]]
    print(f"Headers in Sheet A: {headers_a}")

    # Handle None or invalid headers
    if None in headers_a or not all(headers_a):
        raise ValueError("Sheet A contains empty or invalid column headers.")

    sheet_a_data = [
        {col: cell for col, cell in zip(headers_a, row)}
        for row in sheet_a.iter_rows(min_row=2, values_only=True)
    ]

    # Load Sheet B
    wb_b = openpyxl.load_workbook(sheet_b_path)
    sheet_b = wb_b.active

    # Debug: Print headers
    headers_b = [cell.value for cell in sheet_b[1]]
    print(f"Headers in Sheet B: {headers_b}")

    # Handle None or invalid headers
    if None in headers_b or not all(headers_b):
        raise ValueError("Sheet B contains empty or invalid column headers.")

    sheet_b_data = [
        {col: cell for col, cell in zip(headers_b, row)}
        for row in sheet_b.iter_rows(min_row=2, values_only=True)
    ]

    # Create a dictionary from Sheet B for easy lookup
    sheet_b_lookup = {
        (row['standard'], row['standard number']): row
        for row in sheet_b_data
    }

    # Prepare output data
    output_data = []
    for row_a in sheet_a_data:
        key = (row_a['standard'], row_a['standard number'])
        matched_row_b = sheet_b_lookup.get(key, {})

        # Merge data from Sheet A and Sheet B
        output_row = {
            'standard': row_a['standard'],
            'standard number': row_a['standard number'],
            'category': matched_row_b.get('category', ''),
            'test cases': matched_row_b.get('test cases', ''),
            'expectation': matched_row_b.get('expectation', '')
        }
        output_data.append(output_row)

    # Write to output Excel file
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.append(['standard', 'standard number', 'category', 'test cases', 'expectation'])

    for row in output_data:
        ws_output.append([
            row['standard'],
            row['standard number'],
            row['category'],
            row['test cases'],
            row['expectation']
        ])

    wb_output.save(output_path)

# Paths to input and output files
sheet_a_path = 'sheet-A.xlsx'  # Replace with your actual Sheet A path
sheet_b_path = 'sheet-B.xlsx'  # Replace with your actual Sheet B path
output_path = 'merged_sheet.xlsx'  # Replace with your desired output path

# Call the function to merge sheets
merge_sheets(sheet_a_path, sheet_b_path, output_path)

print(f"Merged data saved to {output_path}")
